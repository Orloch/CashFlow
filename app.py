import io
import datetime
import unicodedata

import streamlit as st
import pandas as pd

# ────────────────────────────────────────────────────────────
def normalize_columns(df):
    """Rimuove accenti, fa lower, underscore al posto degli spazi."""
    clean = []
    for col in df.columns:
        nfkd = unicodedata.normalize("NFKD", col)
        no_acc = "".join(ch for ch in nfkd if not unicodedata.combining(ch))
        clean.append(no_acc.lower().strip().replace(" ", "_"))
    df.columns = clean

def style_with_thousands(df):
    """Formatta tutte le colonne numeriche con separatore di migliaia."""
    num_cols = df.select_dtypes(include="number").columns
    fmt = {col: "{:,.0f}" for col in num_cols}
    return df.style.format(fmt)
# ────────────────────────────────────────────────────────────

st.set_page_config(page_title="Flujo de Caja", layout="wide")
st.title("Monitoreo del Flujo de Caja")

# 1) Upload dei file
prov_file = st.file_uploader("Proveedores (Excel)", type="xlsx")
cli_file  = st.file_uploader("Clientes  (Excel)",    type="xlsx")

if prov_file and cli_file:
    # 2) Lettura e normalizzazione colonne
    df_prov = pd.read_excel(prov_file, engine="openpyxl")
    df_cli  = pd.read_excel(cli_file,  engine="openpyxl")
    normalize_columns(df_prov)
    normalize_columns(df_cli)

    # 3) Copie “raw” per Excel crudo
    df_prov_raw = df_prov.copy()
    df_cli_raw  = df_cli.copy()
    for df in (df_prov_raw, df_cli_raw):
        df["vencimiento_real" if "vencimiento_real" in df.columns else "fecha_de_cobro"] = pd.to_datetime(
            df.get("vencimiento_real", df.get("fecha_de_cobro")), errors="coerce"
        ).dt.date

    # 4) Validazione colonne obbligatorie
    req_prov = {"nombre_proveedor", "valor", "vencimiento_real", "moneda"}
    req_cli  = {"cliente", "total_a_cobrar", "fecha_de_cobro",  "moneda"}
    if not req_prov.issubset(df_prov.columns):
        st.error(f"Faltan columnas en proveedores: {req_prov}")
        st.stop()
    if not req_cli.issubset(df_cli.columns):
        st.error(f"Faltan columnas en clientes:    {req_cli}")
        st.stop()

    # 5) Conversione campi data a datetime.date
    df_prov["vencimiento_real"] = pd.to_datetime(
        df_prov["vencimiento_real"], errors="coerce"
    ).dt.date
    df_cli["fecha_de_cobro"] = pd.to_datetime(
        df_cli["fecha_de_cobro"], errors="coerce"
    ).dt.date

    # 6) Range date e tasso USD affiancati
    min_d = min(df_prov["vencimiento_real"].min(),
                df_cli["fecha_de_cobro"].min())
    max_d = max(df_prov["vencimiento_real"].max(),
                df_cli["fecha_de_cobro"].max())

    col_date, col_rate = st.columns([3, 1])
    with col_date:
        start_date, end_date = st.date_input(
            "Rango de fechas",
            value=(min_d, max_d),
            format="DD-MM-YYYY"
        )
    with col_rate:
        rate = st.number_input(
            "Tasa USD→Local",
            min_value=0.0, value=1.0, format="%.4f"
        )

    # 7) Filtro per intervallo di date
    df_prov = df_prov[
        df_prov["vencimiento_real"].between(start_date, end_date)
    ].copy()
    df_cli = df_cli[
        df_cli["fecha_de_cobro"].between(start_date, end_date)
    ].copy()

    # 8) Arrotondamento importi
    df_prov["valor"]         = df_prov["valor"].round(0).astype(int)
    df_cli["total_a_cobrar"] = df_cli["total_a_cobrar"].round(0).astype(int)

    # 9) Individua dinamicamente le colonne importo
    prov_amt_col = next((c for c in df_prov.columns if "valor" in c), None)
    cli_amt_col  = next((c for c in df_cli.columns  if "total" in c and "cobrar" in c), None)
    if not prov_amt_col or not cli_amt_col:
        st.error("Non ho trovato le colonne importo corrette.")
        st.stop()

    # 10) Applica tasso USD → monto_local
    usd_keys = {"usd","u$s","us$","dolar","dólar","dolares","dólares"}
    def apply_rate(row, amt_col):
        m = str(row["moneda"]).lower()
        val = row[amt_col]
        return round(val * rate) if m in usd_keys else round(val)

    df_prov["monto_local"] = df_prov.apply(lambda r: apply_rate(r, prov_amt_col), axis=1)
    df_cli["monto_local"]  = df_cli.apply(lambda r: apply_rate(r, cli_amt_col),  axis=1)

    # 11) Preview dati con date in DD-MM-YYYY
    preview_prov = df_prov.copy()
    preview_prov["vencimiento_real"] = preview_prov[
        "vencimiento_real"
    ].apply(lambda d: d.strftime("%d-%m-%Y") if d else "")

    preview_cli = df_cli.copy()
    preview_cli["fecha_de_cobro"] = preview_cli[
        "fecha_de_cobro"
    ].apply(lambda d: d.strftime("%d-%m-%Y") if d else "")

    st.subheader("Proveedores (vista previa + tasso)")
    st.dataframe(style_with_thousands(preview_prov), use_container_width=True)
    st.subheader("Clientes (vista previa + tasso)")
    st.dataframe(style_with_thousands(preview_cli),  use_container_width=True)

    # 12) Pivot e consolidato
    pivot_prov = (
        df_prov.groupby("vencimiento_real")["monto_local"].sum()
               .reset_index()
               .pivot_table(index=[], columns="vencimiento_real",
                            values="monto_local", fill_value=0)
    )
    pivot_prov.index = ["Proveedores"]

    pivot_cli = (
        df_cli.groupby("fecha_de_cobro")["monto_local"].sum()
              .reset_index()
              .pivot_table(index=[], columns="fecha_de_cobro",
                           values="monto_local", fill_value=0)
    )
    pivot_cli.index = ["Clientes"]

    df_cassa = pd.concat([pivot_prov, pivot_cli])
    df_cassa = df_cassa.reindex(sorted(df_cassa.columns), axis=1)

    df_cassa.columns = [
        c.strftime("%d-%m-%Y") if isinstance(c, (datetime.date, pd.Timestamp)) else c
        for c in df_cassa.columns
    ]

    st.subheader("Flujo de Caja Consolidado")
    st.dataframe(style_with_thousands(df_cassa), use_container_width=True)

    # 13) Download Excel “crudo” (< start_date)
    df_raw_prov = df_prov_raw[
        df_prov_raw["vencimiento_real"] < start_date
    ]
    df_raw_cli  = df_cli_raw[
        df_cli_raw["fecha_de_cobro"] < start_date
    ]
    buf_raw = io.BytesIO()
    with pd.ExcelWriter(buf_raw, engine="openpyxl") as w:
        df_raw_prov.to_excel(w, "Proveedores", index=False)
        df_raw_cli.to_excel(w, "Clientes",    index=False)
    raw_data = buf_raw.getvalue()

    # 14) Download Excel consolidato
    buf_cons = io.BytesIO()
    with pd.ExcelWriter(buf_cons, engine="openpyxl") as w:
        df_cassa.to_excel(w, "Consolidado")
    cons_data = buf_cons.getvalue()

    st.subheader("Proveedores (rango + cambio)")
    st.dataframe(style_with_thousands(df_prov), use_container_width=True)
    st.subheader("Clientes (rango + cambio)")
    st.dataframe(style_with_thousands(df_cli),  use_container_width=True)

    # 10) Pivot & concatenazione
    pivot_prov = (
        df_prov
        .groupby("vencimiento_real")["monto_local"]
        .sum()
        .reset_index()
        .pivot_table(index=[], columns="vencimiento_real",
                     values="monto_local", fill_value=0)
    )
    pivot_prov.index = ["Proveedores"]

    pivot_cli = (
        df_cli
        .groupby("fecha_de_cobro")["monto_local"]
        .sum()
        .reset_index()
        .pivot_table(index=[], columns="fecha_de_cobro",
                     values="monto_local", fill_value=0)
    )
    pivot_cli.index = ["Clientes"]

    df_cassa = pd.concat([pivot_prov, pivot_cli])
    df_cassa = df_cassa.reindex(sorted(df_cassa.columns), axis=1)

    # 11) Rinomina intestazioni date in DD-MM-YYYY
    df_cassa.columns = [
        c.strftime("%d-%m-%Y") if isinstance(c, (datetime.date, pd.Timestamp))
        else c
        for c in df_cassa.columns
    ]

    # 12) Mostra consolidato
    st.subheader("Flujo de Caja Consolidado")
    st.dataframe(style_with_thousands(df_cassa), use_container_width=True)

    # 13) Report crudo: solo righe con vencimiento_real < start_date
    raw_prov = df_prov_raw[
        df_prov_raw["vencimiento_real"] < start_date
    ].copy()
    raw_cli = df_cli_raw[
        df_cli_raw["fecha_de_cobro"] < start_date
    ].copy()

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo consolidado"

    buf_raw = io.BytesIO()
    with pd.ExcelWriter(buf_cons, engine="openpyxl") as w:
        df_cassa.to_excel(w, sheet_name="Resumen", index=True)
        df_prov.to_excel(w, sheet_name="Detalle Proveedores", index=False)
        df_cli.to_excel(w, sheet_name="Detalle Clientes", index=False)
    cons_data = buf_cons.getvalue()

    # 14) Report consolidato
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    # 1) Fornitori con prefisso "PROV -"
    prov_tab = df_prov[["nombre_proveedor", "vencimiento_real", "monto_local"]].copy()
    prov_tab["Entidad"] = "PROV - " + prov_tab["nombre_proveedor"]
    prov_tab["Fecha"] = prov_tab["vencimiento_real"].apply(lambda d: d.strftime("%d-%m-%Y") if pd.notnull(d) else "")
    prov_tab["Importo"] = -prov_tab["monto_local"]  # val. negativi
    prov_tab = prov_tab[["Entidad", "Fecha", "Importo"]]

    # 2) Clienti con prefisso "CLI -"
    cli_tab = df_cli[["cliente", "fecha_de_cobro", "monto_local"]].copy()
    cli_tab["Entidad"] = "CLI - " + cli_tab["cliente"]
    cli_tab["Fecha"] = cli_tab["fecha_de_cobro"].apply(lambda d: d.strftime("%d-%m-%Y") if pd.notnull(d) else "")
    cli_tab["Importo"] = cli_tab["monto_local"]
    cli_tab = cli_tab[["Entidad", "Fecha", "Importo"]]

    # 3) Unione e pivot
    union_tab = pd.concat([prov_tab, cli_tab], ignore_index=True)
    pivot = union_tab.pivot_table(
        index="Entidad",
        columns="Fecha",
        values="Importo",
        aggfunc="sum",
        fill_value=0
    )

    # 4) Ordine manuale: prima PROV poi CLI
    new_order = sorted([i for i in pivot.index if i.startswith("PROV")]) + \
                sorted([i for i in pivot.index if i.startswith("CLI")])
    pivot = pivot.loc[new_order]

    # 5) Riga totale alla fine
    total_row = pd.DataFrame(pivot.sum(axis=0)).T
    total_row.index = ["TOTAL FLUJO"]
    pivot = pd.concat([pivot, total_row])

    for r_idx, row in enumerate(dataframe_to_rows(pivot, index=True, header=True), start=1):
        cleaned_row = []
        for val in row:
            if pd.isna(val):
                cleaned_row.append("")
            elif isinstance(val, (pd.Timestamp, datetime.date)):
                cleaned_row.append(val.strftime("%d-%m-%Y"))
            elif isinstance(val, (int, float, str)):
                cleaned_row.append(val)
            else:
                cleaned_row.append(str(val))
        ws.append(cleaned_row)

        is_total_row = (ws.cell(row=r_idx, column=1).value == "TOTAL FLUJO")

        for c_idx, cell in enumerate(ws[r_idx], start=1):
            # Intestazioni (prima riga)
            if r_idx == 1:
                cell.font = Font(bold=False)
                cell.alignment = Alignment(horizontal="center")

            # Riga totale → solo questa va in grassetto
            elif is_total_row:
                cell.font = Font(bold=True)
                if c_idx > 1:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")

            # Tutte le altre righe normali
            else:
                cell.font = Font(bold=False)
                if c_idx > 1:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
    git - -version
    # Esporta
    buf_final = io.BytesIO()
    wb.save(buf_final)

    # 15) Bottoni di download
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "Descargar CRUDO (< start_date)",
            raw_data,  # oppure raw_bytes se l’hai rinominato così
            "datos_crudos.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_raw_excel"
        )
    with c2:
        st.download_button(
            "Descargar Flujo consolidado (horizontal)",
            buf_final.getvalue(),
            "flujo_consolidado_horizontal.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="flujo_horizontal_excel"
        )
